import asyncio
import io
import json
import os
import secrets
import zipfile
from datetime import datetime, timedelta, timezone
from typing import Any, Dict

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from cryptography.fernet import Fernet, InvalidToken
from fastapi import HTTPException
from fastapi.responses import RedirectResponse
from google.auth.transport.requests import Request as GoogleRequest
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import re

SCOPES = [
    "openid",
    "email",
    "https://www.googleapis.com/auth/drive.file",
]
_scheduler = None


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def client_config():
    client_id = os.environ.get("GOOGLE_DRIVE_CLIENT_ID") or os.environ.get("GOOGLE_CLIENT_ID")
    secret = os.environ.get("GOOGLE_DRIVE_CLIENT_SECRET")
    redirect = os.environ.get("GOOGLE_DRIVE_REDIRECT_URI")
    if not client_id or not secret or not redirect:
        raise HTTPException(status_code=500, detail="Google Drive OAuth is not configured")
    return {
        "web": {
            "client_id": client_id,
            "client_secret": secret,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [redirect],
        }
    }


def cipher():
    key = (
        os.environ.get("BACKUP_ENCRYPTION_KEY", "").strip()
        or os.environ.get("BACKUP_TOKEN_ENCRYPTION_KEY", "").strip()
    )
    try:
        return Fernet(key.encode())
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Invalid backup encryption key") from exc


def encrypt_credentials(credentials):
    payload = {
        "token": credentials.token,
        "refresh_token": credentials.refresh_token,
        "token_uri": credentials.token_uri,
        "client_id": credentials.client_id,
        "client_secret": credentials.client_secret,
        "scopes": list(credentials.scopes or SCOPES),
    }
    return cipher().encrypt(json.dumps(payload).encode()).decode()


def load_credentials(document):
    try:
        payload = json.loads(
            cipher().decrypt(document["encrypted_credentials"].encode()).decode()
        )
    except (InvalidToken, ValueError, KeyError) as exc:
        raise HTTPException(status_code=500, detail="Stored Drive authorization is unreadable") from exc

    credentials = Credentials(**payload)
    if credentials.expired and credentials.refresh_token:
        credentials.refresh(GoogleRequest())
    return credentials


def json_safe(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if value.__class__.__name__ == "ObjectId":
        return str(value)
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    return value


async def create_archive(db):
    excluded = {"user_sessions", "backup_oauth_states"}
    names = [
        name for name in await db.list_collection_names()
        if name not in excluded and not name.startswith("system.")
    ]
    collections = {}
    counts = {}
    total = 0

    for name in sorted(names):
        documents = await db[name].find({}).to_list(length=None)
        safe = [json_safe(document) for document in documents]
        collections[name] = safe
        counts[name] = len(safe)
        total += len(safe)

    metadata = {
        "application": "MDS Laboratory Information Management System",
        "version": os.environ.get("APP_VERSION", "2.2"),
        "created_at_utc": now_iso(),
        "database_name": os.environ.get("DB_NAME", ""),
        "collection_count": len(names),
        "total_documents": total,
        "collection_counts": counts,
        "excluded_collections": sorted(excluded),
    }

    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("backup.json", json.dumps({"collections": collections}, indent=2))
        archive.writestr("metadata.json", json.dumps(metadata, indent=2))
        archive.writestr("version.txt", metadata["version"])
    return stream.getvalue(), metadata, collections



def excel_scalar(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    return json.dumps(json_safe(value), ensure_ascii=False)


def safe_sheet_name(name, used):
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", str(name or "Sheet")).strip()
    cleaned = cleaned[:31] or "Sheet"
    candidate = cleaned
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def flatten_document(document, prefix=""):
    result = {}
    for key, value in document.items():
        if key == "_id":
            continue

        column = f"{prefix}.{key}" if prefix else str(key)

        if isinstance(value, dict):
            result.update(flatten_document(value, column))
        elif isinstance(value, list):
            # Lists remain readable as JSON in one cell rather than creating
            # unpredictable columns.
            result[column] = json.dumps(
                json_safe(value),
                ensure_ascii=False,
            )
        else:
            result[column] = excel_scalar(value)
    return result


def style_worksheet(worksheet, freeze="A2"):
    worksheet.freeze_panes = freeze
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        maximum = 0
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            maximum = max(maximum, len(value))
        worksheet.column_dimensions[letter].width = min(
            max(maximum + 2, 10),
            38,
        )


def create_human_readable_excel(collections, metadata):
    """
    Build an Excel companion backup intended for people to inspect.

    - Summary: backup metadata and collection counts.
    - Laboratory Records: one row per test with patient/sample/result context.
    - Other collections: one sheet per collection with flattened documents.

    The Excel workbook is informational only. The ZIP/JSON archive remains
    the authoritative restore source.
    """
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    used_sheet_names = set()

    summary = workbook.create_sheet(
        safe_sheet_name("Summary", used_sheet_names)
    )
    summary_rows = [
        ["MDS LIMS Human-readable Backup", ""],
        ["Created at (UTC)", metadata.get("created_at_utc", "")],
        ["Application", metadata.get("application", "")],
        ["Application Version", metadata.get("version", "")],
        ["Database", metadata.get("database_name", "")],
        ["Collections", metadata.get("collection_count", 0)],
        ["Total Documents", metadata.get("total_documents", 0)],
        ["", ""],
        ["Collection", "Documents"],
    ]
    for name, count in sorted(
        (metadata.get("collection_counts") or {}).items()
    ):
        summary_rows.append([name, count])

    for row in summary_rows:
        summary.append(row)

    summary["A1"].font = Font(bold=True, size=16)
    summary["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="B7DEE8",
    )
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 50
    summary.freeze_panes = "A10"

    records = collections.get("records") or []
    records_sheet = workbook.create_sheet(
        safe_sheet_name("Laboratory Records", used_sheet_names)
    )

    record_headers = [
        "Patient ID",
        "Patient Name",
        "Age",
        "Sex",
        "District",
        "Requesting Institution",
        "Date Received",
        "Patient EPID",
        "Dataset",
        "Lab Number",
        "Sample Type",
        "NVHCP Programme",
        "Assigned Panels",
        "Test",
        "Result",
        "Additional Result",
        "Result Date",
        "Test Remarks",
        "Sample Remarks",
    ]
    records_sheet.append(record_headers)

    for record in records:
        samples = record.get("samples") or []

        # Legacy records without nested samples are still represented.
        if not samples:
            tests = record.get("tests") or []
            if not tests:
                tests = [{}]

            for test in tests:
                records_sheet.append(
                    [
                        record.get("id", ""),
                        record.get("name", ""),
                        record.get("age", ""),
                        record.get("sex", ""),
                        record.get("district", ""),
                        record.get("requesting_institution", ""),
                        record.get("date", ""),
                        record.get("epid_number", ""),
                        record.get("dataset", ""),
                        record.get("lab_number", ""),
                        record.get("sample_type", ""),
                        "",
                        "",
                        test.get("test", ""),
                        test.get("result1", ""),
                        test.get("result2", ""),
                        test.get("result_date", ""),
                        test.get("remarks", ""),
                        record.get("remarks", ""),
                    ]
                )
            continue

        for sample in samples:
            panel_names = ", ".join(
                str(panel.get("panel_name") or "").strip()
                for panel in (sample.get("assigned_panels") or [])
                if str(panel.get("panel_name") or "").strip()
            )

            tests = sample.get("tests") or []
            if not tests:
                tests = [{}]

            for test in tests:
                records_sheet.append(
                    [
                        record.get("id", ""),
                        record.get("name", ""),
                        record.get("age", ""),
                        record.get("sex", ""),
                        record.get("district", ""),
                        record.get("requesting_institution", ""),
                        record.get("date", ""),
                        record.get("epid_number", ""),
                        sample.get("dataset")
                        or record.get("dataset", ""),
                        sample.get("lab_number", ""),
                        sample.get("sample_type", ""),
                        "Yes"
                        if sample.get("nvhcp_program")
                        else "No",
                        panel_names,
                        test.get("test", ""),
                        test.get("result1", ""),
                        test.get("result2", ""),
                        test.get("result_date", ""),
                        test.get("remarks", ""),
                        sample.get("remarks", ""),
                    ]
                )

    style_worksheet(records_sheet)

    # Keep patient and result columns comfortably readable.
    preferred_widths = {
        "B": 24,
        "E": 18,
        "F": 28,
        "I": 20,
        "J": 18,
        "K": 20,
        "M": 28,
        "N": 32,
        "O": 20,
        "P": 24,
        "R": 30,
        "S": 30,
    }
    for column, width in preferred_widths.items():
        records_sheet.column_dimensions[column].width = width

    # Add all non-record collections as generic readable sheets.
    for collection_name, documents in sorted(collections.items()):
        if collection_name == "records":
            continue

        worksheet = workbook.create_sheet(
            safe_sheet_name(collection_name, used_sheet_names)
        )

        flattened = [
            flatten_document(document)
            for document in documents
        ]
        columns = sorted(
            {
                key
                for document in flattened
                for key in document.keys()
            }
        )

        if not columns:
            worksheet.append(["No data"])
            continue

        worksheet.append(columns)
        for document in flattened:
            worksheet.append(
                [
                    excel_scalar(document.get(column, ""))
                    for column in columns
                ]
            )

        style_worksheet(worksheet)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()



def drive_service(credentials):
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def folder_id(service):
    configured = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "").strip()
    if configured:
        return configured

    name = os.environ.get("GOOGLE_DRIVE_FOLDER_NAME", "MDS LIMS Backups")
    escaped = name.replace("'", "\\'")
    found = service.files().list(
        q=f"name='{escaped}' and mimeType='application/vnd.google-apps.folder' and trashed=false",
        spaces="drive",
        fields="files(id,name)",
    ).execute().get("files", [])
    if found:
        return found[0]["id"]

    return service.files().create(
        body={"name": name, "mimeType": "application/vnd.google-apps.folder"},
        fields="id",
    ).execute()["id"]


def upload(
    credentials,
    data,
    filename,
    mimetype="application/octet-stream",
):
    service = drive_service(credentials)
    folder = folder_id(service)
    media = MediaIoBaseUpload(
        io.BytesIO(data),
        mimetype=mimetype,
        resumable=False,
    )
    created = service.files().create(
        body={"name": filename, "parents": [folder]},
        media_body=media,
        fields="id,name,size,createdTime,webViewLink",
    ).execute()
    return {"folder_id": folder, **created}


def list_files(credentials, limit=30):
    service = drive_service(credentials)
    folder = folder_id(service)
    files = service.files().list(
        q=f"'{folder}' in parents and trashed=false and name contains 'MDS_LIMS_Backup_'",
        orderBy="createdTime desc",
        pageSize=min(max(limit * 2, 1), 200),
        fields="files(id,name,size,createdTime,webViewLink)",
    ).execute().get("files", [])

    return [
        item
        for item in files
        if str(item.get("name") or "").lower().endswith(".zip")
    ][:limit]


def apply_retention(credentials):
    keep = int(os.environ.get("BACKUP_RETENTION_COUNT", "30"))
    service = drive_service(credentials)
    folder = folder_id(service)

    files = service.files().list(
        q=f"'{folder}' in parents and trashed=false",
        orderBy="createdTime desc",
        pageSize=1000,
        fields="files(id,name,createdTime)",
    ).execute().get("files", [])

    zip_files = [
        item
        for item in files
        if str(item.get("name") or "").startswith("MDS_LIMS_Backup_")
        and str(item.get("name") or "").lower().endswith(".zip")
    ]

    deleted = 0
    old_zip_files = zip_files[keep:]

    for zip_file in old_zip_files:
        name = zip_file.get("name") or ""
        timestamp = name[
            len("MDS_LIMS_Backup_") : -len(".zip")
        ]

        service.files().delete(
            fileId=zip_file["id"]
        ).execute()
        deleted += 1

        companion_name = f"MDS_LIMS_Data_{timestamp}.xlsx"
        for item in files:
            if item.get("name") == companion_name:
                try:
                    service.files().delete(
                        fileId=item["id"]
                    ).execute()
                    deleted += 1
                except Exception:
                    pass

    return deleted


async def perform_backup(db, trigger, apply_cleanup=True):
    document = await db.backup_settings.find_one({"key": "google_drive"}, {"_id": 0})
    if not document or not document.get("encrypted_credentials"):
        raise HTTPException(status_code=409, detail="Google Drive is not connected")

    credentials = await asyncio.to_thread(load_credentials, document)
    archive, metadata, collections = await create_archive(db)

    timestamp = datetime.now(timezone.utc).strftime(
        "%Y-%m-%d_%H%M%S"
    )
    filename = f"MDS_LIMS_Backup_{timestamp}.zip"
    excel_filename = f"MDS_LIMS_Data_{timestamp}.xlsx"

    excel_data = await asyncio.to_thread(
        create_human_readable_excel,
        collections,
        metadata,
    )

    uploaded = await asyncio.to_thread(
        upload,
        credentials,
        archive,
        filename,
        "application/zip",
    )
    excel_uploaded = await asyncio.to_thread(
        upload,
        credentials,
        excel_data,
        excel_filename,
        (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    deleted = (
        await asyncio.to_thread(apply_retention, credentials)
        if apply_cleanup
        else 0
    )

    update = {
        "last_backup_at": now_iso(),
        "last_backup_status": "success",
        "last_backup_name": filename,
        "last_backup_size": uploaded.get("size") or len(archive),
        "last_excel_backup_name": excel_filename,
        "last_excel_backup_size":
            excel_uploaded.get("size") or len(excel_data),
        "last_backup_trigger": trigger,
        "folder_id": uploaded.get("folder_id"),
        "last_error": None,
        "updated_at": now_iso(),
    }
    await db.backup_settings.update_one(
        {"key": "google_drive"}, {"$set": update}, upsert=True
    )
    await db.backup_logs.insert_one({
        "id": secrets.token_hex(8),
        "created_at": now_iso(),
        "status": "success",
        "trigger": trigger,
        "filename": filename,
        "size": update["last_backup_size"],
        "excel_filename": excel_filename,
        "excel_size": update["last_excel_backup_size"],
        "metadata": metadata,
        "retention_deleted": deleted,
    })
    return {
        "ok": True,
        "filename": filename,
        "file": uploaded,
        "excel_filename": excel_filename,
        "excel_file": excel_uploaded,
        "metadata": metadata,
    }



def get_drive_file(credentials, file_id):
    service = drive_service(credentials)
    return service.files().get(
        fileId=file_id,
        fields="id,name,size,createdTime,parents,trashed",
    ).execute()


def download_drive_file(credentials, file_id, max_bytes=200 * 1024 * 1024):
    service = drive_service(credentials)
    metadata = get_drive_file(credentials, file_id)

    if metadata.get("trashed"):
        raise HTTPException(
            status_code=404,
            detail="The selected backup file is in Google Drive Trash",
        )

    size = int(metadata.get("size") or 0)
    if size and size > max_bytes:
        raise HTTPException(
            status_code=413,
            detail="The selected backup is too large to restore safely",
        )

    request = service.files().get_media(fileId=file_id)
    stream = io.BytesIO()
    downloader = MediaIoBaseDownload(
        stream,
        request,
        chunksize=1024 * 1024,
    )

    done = False
    while not done:
        _, done = downloader.next_chunk()

    data = stream.getvalue()
    if len(data) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail="The selected backup is too large to restore safely",
        )

    return data, metadata


def parse_backup_archive(data):
    try:
        with zipfile.ZipFile(io.BytesIO(data), "r") as archive:
            names = set(archive.namelist())
            if "backup.json" not in names:
                raise HTTPException(
                    status_code=400,
                    detail="Invalid MDS LIMS backup: backup.json is missing",
                )

            payload = json.loads(
                archive.read("backup.json").decode("utf-8")
            )
            metadata = {}
            if "metadata.json" in names:
                metadata = json.loads(
                    archive.read("metadata.json").decode("utf-8")
                )
    except HTTPException:
        raise
    except (zipfile.BadZipFile, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise HTTPException(
            status_code=400,
            detail="The selected file is not a valid MDS LIMS backup",
        ) from exc

    collections = payload.get("collections")
    if not isinstance(collections, dict):
        raise HTTPException(
            status_code=400,
            detail="Invalid MDS LIMS backup structure",
        )

    application = str(metadata.get("application") or "")
    if application and "MDS" not in application.upper():
        raise HTTPException(
            status_code=400,
            detail="The selected archive does not appear to be an MDS LIMS backup",
        )

    total_documents = 0
    safe_collections = {}
    for name, documents in collections.items():
        if (
            not isinstance(name, str)
            or not re.match(r"^[A-Za-z0-9_.-]+$", name)
            or name.startswith("system.")
        ):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid collection name in backup: {name}",
            )
        if not isinstance(documents, list):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid document list for collection: {name}",
            )
        if not all(isinstance(document, dict) for document in documents):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid document in collection: {name}",
            )

        total_documents += len(documents)
        safe_collections[name] = documents

    if total_documents > 1_000_000:
        raise HTTPException(
            status_code=413,
            detail="Backup contains too many documents to restore safely",
        )

    return safe_collections, metadata


async def inspect_backup_from_drive(db, file_id):
    settings = await db.backup_settings.find_one(
        {"key": "google_drive"},
        {"_id": 0},
    )
    if not settings or not settings.get("encrypted_credentials"):
        raise HTTPException(
            status_code=409,
            detail="Google Drive is not connected",
        )

    credentials = await asyncio.to_thread(load_credentials, settings)
    data, drive_metadata = await asyncio.to_thread(
        download_drive_file,
        credentials,
        file_id,
    )
    collections, metadata = parse_backup_archive(data)

    return {
        "file": drive_metadata,
        "metadata": metadata,
        "collection_count": len(collections),
        "total_documents": sum(
            len(documents)
            for documents in collections.values()
        ),
        "collection_counts": {
            name: len(documents)
            for name, documents in sorted(collections.items())
        },
    }


async def restore_backup_from_drive(db, file_id, expected_filename=None):
    """
    Restore collections contained in a selected Drive backup.

    Runtime security/session and Drive-connection collections are preserved so
    the user is not logged out and the active Google Drive authorization is not
    replaced by older credentials.
    """
    settings = await db.backup_settings.find_one(
        {"key": "google_drive"},
        {"_id": 0},
    )
    if not settings or not settings.get("encrypted_credentials"):
        raise HTTPException(
            status_code=409,
            detail="Google Drive is not connected",
        )

    credentials = await asyncio.to_thread(load_credentials, settings)
    data, drive_metadata = await asyncio.to_thread(
        download_drive_file,
        credentials,
        file_id,
    )

    if (
        expected_filename
        and drive_metadata.get("name")
        and drive_metadata["name"] != expected_filename
    ):
        raise HTTPException(
            status_code=409,
            detail="The selected Google Drive backup changed; refresh and try again",
        )

    collections, metadata = parse_backup_archive(data)

    # Create a current-state backup before any destructive operation.
    # Retention cleanup is deliberately skipped here so the selected old
    # backup cannot be removed by the pre-restore backup itself.
    pre_restore = await perform_backup(
        db,
        "pre_restore",
        apply_cleanup=False,
    )

    protected = {
        "user_sessions",
        "master_sessions",
        "backup_oauth_states",
        "backup_settings",
        "backup_logs",
    }

    restored_counts = {}
    try:
        for name, documents in collections.items():
            if name in protected:
                continue

            collection = db[name]
            await collection.delete_many({})

            cleaned = []
            for document in documents:
                item = dict(document)
                # Backups serialize ObjectId as strings. Let MongoDB create
                # fresh internal _id values; MDS uses its own stable id fields.
                item.pop("_id", None)
                cleaned.append(item)

            if cleaned:
                await collection.insert_many(
                    cleaned,
                    ordered=False,
                )

            restored_counts[name] = len(cleaned)

    except Exception as exc:
        await db.backup_logs.insert_one(
            {
                "id": secrets.token_hex(8),
                "created_at": now_iso(),
                "status": "restore_failed",
                "trigger": "restore",
                "source_filename": drive_metadata.get("name"),
                "source_file_id": file_id,
                "pre_restore_backup": pre_restore.get("filename"),
                "error": str(exc),
            }
        )
        raise HTTPException(
            status_code=500,
            detail=(
                "Restore failed after the safety backup was created. "
                f"Pre-restore backup: {pre_restore.get('filename')}"
            ),
        ) from exc

    await db.backup_logs.insert_one(
        {
            "id": secrets.token_hex(8),
            "created_at": now_iso(),
            "status": "restore_success",
            "trigger": "restore",
            "source_filename": drive_metadata.get("name"),
            "source_file_id": file_id,
            "source_metadata": metadata,
            "pre_restore_backup": pre_restore.get("filename"),
            "restored_collections": restored_counts,
        }
    )

    return {
        "ok": True,
        "source_filename": drive_metadata.get("name"),
        "pre_restore_backup": pre_restore.get("filename"),
        "restored_collections": restored_counts,
        "restored_documents": sum(restored_counts.values()),
        "metadata": metadata,
    }


async def scheduled_backup(db):
    if os.environ.get("ENABLE_GOOGLE_DRIVE_BACKUP", "false").lower() != "true":
        return
    try:
        await perform_backup(db, "scheduled")
    except Exception as exc:
        await db.backup_settings.update_one(
            {"key": "google_drive"},
            {"$set": {
                "last_backup_at": now_iso(),
                "last_backup_status": "failed",
                "last_error": str(exc),
                "updated_at": now_iso(),
            }},
            upsert=True,
        )


def register_backup(app, router, db, get_current_user, require_master_access):
    from fastapi import Depends, Request
    from apscheduler.triggers.cron import CronTrigger

    @router.get("/backup/google/connect-url")
    async def connect_url(user=Depends(require_master_access)):
        state = secrets.token_urlsafe(32)
        await db.backup_oauth_states.insert_one({
            "state": state,
            "user_id": user["user_id"],
            "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat(),
        })
        flow = Flow.from_client_config(
            client_config(),
            scopes=SCOPES,
            redirect_uri=os.environ["GOOGLE_DRIVE_REDIRECT_URI"],
        )
        url, _ = flow.authorization_url(
            access_type="offline",
            include_granted_scopes="true",
            prompt="consent",
            state=state,
        )
        return {"authorization_url": url}

    @router.get("/backup/google/callback")
    async def callback(request: Request, code: str, state: str):
        state_doc = await db.backup_oauth_states.find_one({"state": state}, {"_id": 0})
        if not state_doc:
            raise HTTPException(status_code=400, detail="Invalid OAuth state")

        expires = datetime.fromisoformat(state_doc["expires_at"])
        if expires.tzinfo is None:
            expires = expires.replace(tzinfo=timezone.utc)
        if expires < datetime.now(timezone.utc):
            raise HTTPException(status_code=400, detail="OAuth state expired")

        flow = Flow.from_client_config(
            client_config(),
            scopes=SCOPES,
            redirect_uri=os.environ["GOOGLE_DRIVE_REDIRECT_URI"],
            state=state,
        )
        flow.fetch_token(code=code)
        credentials = flow.credentials

        previous = await db.backup_settings.find_one({"key": "google_drive"}, {"_id": 0})
        if not credentials.refresh_token and previous:
            old = load_credentials(previous)
            credentials.refresh_token = old.refresh_token
        if not credentials.refresh_token:
            raise HTTPException(status_code=400, detail="Google did not return a refresh token")

        await db.backup_settings.update_one(
            {"key": "google_drive"},
            {"$set": {
                "key": "google_drive",
                "connected": True,
                "connected_by_user_id": state_doc["user_id"],
                "encrypted_credentials": encrypt_credentials(credentials),
                "connected_at": now_iso(),
                "updated_at": now_iso(),
            }},
            upsert=True,
        )
        await db.backup_oauth_states.delete_one({"state": state})
        frontend = os.environ.get("FRONTEND_URL", "").rstrip("/")
        return RedirectResponse(f"{frontend}/backup?drive=connected")

    @router.get("/backup/status")
    async def status(user=Depends(require_master_access)):
        document = await db.backup_settings.find_one(
            {"key": "google_drive"},
            {"_id": 0, "encrypted_credentials": 0},
        )
        return {
            "connected": bool(document and document.get("connected")),
            "automatic_enabled": os.environ.get(
                "ENABLE_GOOGLE_DRIVE_BACKUP", "false"
            ).lower() == "true",
            "schedule": (
                f"{int(os.environ.get('BACKUP_HOUR', '2')):02d}:"
                f"{int(os.environ.get('BACKUP_MINUTE', '0')):02d}"
            ),
            "timezone": os.environ.get("BACKUP_TIMEZONE", "Asia/Kolkata"),
            "retention_count": int(os.environ.get("BACKUP_RETENTION_COUNT", "30")),
            **(document or {}),
        }

    @router.get("/backup/history")
    async def history(limit: int = 30, user=Depends(require_master_access)):
        document = await db.backup_settings.find_one({"key": "google_drive"}, {"_id": 0})
        if not document:
            return {"items": []}
        credentials = await asyncio.to_thread(load_credentials, document)
        files = await asyncio.to_thread(list_files, credentials, limit)
        return {"items": files}

    @router.post("/backup/run")
    async def run(user=Depends(require_master_access)):
        return await perform_backup(db, "manual")

    @router.get("/backup/inspect/{file_id}")
    async def inspect_backup(
        file_id: str,
        user=Depends(require_master_access),
    ):
        return await inspect_backup_from_drive(db, file_id)

    @router.post("/backup/restore")
    async def restore_backup(
        payload: Dict[str, Any],
        user=Depends(require_master_access),
    ):
        if str(payload.get("confirmation") or "").strip() != "RESTORE":
            raise HTTPException(
                status_code=400,
                detail='Confirmation must be exactly "RESTORE"',
            )

        file_id = str(payload.get("file_id") or "").strip()
        filename = str(payload.get("filename") or "").strip() or None
        if not file_id:
            raise HTTPException(
                status_code=400,
                detail="Backup file is required",
            )

        return await restore_backup_from_drive(
            db,
            file_id,
            filename,
        )

    @router.post("/backup/disconnect")
    async def disconnect(user=Depends(require_master_access)):
        await db.backup_settings.delete_one({"key": "google_drive"})
        return {"ok": True}

    @app.on_event("startup")
    async def start_scheduler():
        global _scheduler
        if _scheduler:
            return
        timezone_name = os.environ.get("BACKUP_TIMEZONE", "Asia/Kolkata")
        _scheduler = AsyncIOScheduler(timezone=timezone_name)
        _scheduler.add_job(
            scheduled_backup,
            CronTrigger(
                hour=int(os.environ.get("BACKUP_HOUR", "2")),
                minute=int(os.environ.get("BACKUP_MINUTE", "0")),
                timezone=timezone_name,
            ),
            args=[db],
            id="mds_drive_backup",
            replace_existing=True,
            max_instances=1,
            coalesce=True,
        )
        _scheduler.start()

    @app.on_event("shutdown")
    async def stop_scheduler():
        global _scheduler
        if _scheduler:
            _scheduler.shutdown(wait=False)
            _scheduler = None
